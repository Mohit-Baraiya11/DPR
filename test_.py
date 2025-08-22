import sys
import os
import psutil
import threading
import datetime
import json
import pickle
from typing import List, Optional
import time
import signal
import subprocess
import atexit

# PyQt5 imports
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QWidget, 
                             QLineEdit, QPushButton, QLabel, QHBoxLayout, 
                             QMessageBox, QFrame, QDialog, QFormLayout,
                             QFileDialog, QTextEdit, QDialogButtonBox, QScrollArea,
                             QTabWidget, QGridLayout, QSpacerItem, QSizePolicy,
                             QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView,
                             QCheckBox)
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QTimer, Qt, pyqtSignal, QObject
from PyQt5.QtGui import QFont, QPalette, QColor

# Server related imports
import openpyxl
import uvicorn
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

# Ngrok import
import pyngrok.ngrok as ngrok
import re

# AI related imports
try:
    from agno.agent import Agent
    from agno.models.groq import Groq
    AGNO_AVAILABLE = True
except ImportError:
    AGNO_AVAILABLE = False
    print("Warning: agno package not available. AI features will be disabled.")

# Configuration storage file
CONFIG_FILE = "app_config.pkl"

# Global variables for server
app_config = {}
fastapi_app = None

# ======================== CONFIGURATION MANAGEMENT ========================

def save_config(config):
    """Save configuration to pickle file"""
    try:
        with open(CONFIG_FILE, 'wb') as f:
            pickle.dump(config, f)
        return True
    except Exception as e:
        print(f"Error saving config: {e}")
        return False

def load_config():
    """Load configuration from pickle file"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'rb') as f:
                return pickle.load(f)
    except Exception as e:
        print(f"Error loading config: {e}")
    return {}

# ======================== SERVER COMPONENTS ========================

# Define the SupportResult model


# Utility functions
def get_available_sheets(file_path: str) -> List[str]:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = [sheet.strip() for sheet in wb.sheetnames if sheet.strip() != "LOGS"]
        wb.close()
        return sheets
    except FileNotFoundError as e:
        raise FileNotFoundError(f"The specified file was not found: {file_path}") from e
    except Exception as e:
        raise Exception(f"Failed to read sheets from {file_path}: {str(e)}") from e

def get_descriptions_with_index(file_path, sheet_name="July.25"):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    descriptions = []
    for row_num in range(5, ws.max_row + 1):
        cell_value = ws[f"C{row_num}"].value
        if cell_value is not None and str(cell_value).strip() != "":
            descriptions.append((row_num, cell_value))
    wb.close()
    return str(descriptions)

def get_date_column(file_path, sheet_name="July.25", date: datetime.date = None):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    column = None
    for cell in ws[1]:
        if isinstance(cell.value, datetime.datetime):
            if date is None:
                if cell.value.date() == datetime.date.today():
                    column = cell.column + 1
                    break
            else:
                if cell.value.date() == date:
                    column = cell.column + 1
                    break
    wb.close()
    return column

def put_logs_in_file(file_path: str, sheet_name="LOGS", description=None, found_description=None,
                     row_index=None, column_index=None, value: float = None, name: str = None,
                     location: str = None, remark: str = None):
    wb = openpyxl.load_workbook(file_path)
    if "LOGS" not in wb.sheetnames:
        ws = wb.create_sheet("LOGS")
        headers = ["Logged_At", "Updated_Sheet", "Name", "Location", "User Description",
                   "Matched Description", "Row", "Column", "Value", "Remark"]
        ws.append(headers)
    else:
        ws = wb["LOGS"]
    next_row = ws.max_row + 1
    while ws.cell(row=next_row, column=1).value not in (None, ""):
        next_row += 1
    ws.cell(row=next_row, column=1, value=datetime.datetime.now())
    ws.cell(row=next_row, column=2, value=sheet_name)
    ws.cell(row=next_row, column=3, value=name)
    ws.cell(row=next_row, column=4, value=location)
    ws.cell(row=next_row, column=5, value=description)
    ws.cell(row=next_row, column=6, value=found_description)
    ws.cell(row=next_row, column=7, value=row_index)
    ws.cell(row=next_row, column=8, value=column_index)
    ws.cell(row=next_row, column=9, value=value)
    ws.cell(row=next_row, column=10, value=remark)
    wb.save(file_path)
    wb.close()

def update_sheet(file_path: str, sheet_name: str = "July.25", row_index: int = None,
                 column_index: int = None, value: float = None) -> None:
    if row_index is None or column_index is None or value is None:
        raise ValueError("row_index, column_index, and value must be provided")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        cell = ws.cell(row=row_index, column=column_index)
        current_value = 0
        if cell.value is not None:
            try:
                current_value = float(cell.value)
            except (ValueError, TypeError):
                print(f"Existing value '{cell.value}' in cell {row_index},{column_index} is not a number. Treating as 0.")
        new_value = current_value + value
        cell.value = new_value
        wb.save(file_path)
        wb.close()
    except Exception as e:
        raise

def get_history(file_path: str, name: str, location: str):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb["LOGS"]
        remarks = []
        for row_num in range(1, ws.max_row + 1):
            user_name = ws[f"C{row_num}"].value
            user_location = ws[f"D{row_num}"].value
            if user_name == name and user_location == location:
                remarks.append({
                    "data": ws[f"A{row_num}"].value,
                    "name": user_name,
                    "location": user_location,
                    "description": ws[f"E{row_num}"].value,
                    "found_description": ws[f"F{row_num}"].value,
                    "row": ws[f"G{row_num}"].value,
                    "column": ws[f"H{row_num}"].value,
                    "value": ws[f"I{row_num}"].value,
                    "conclution": ws[f"J{row_num}"].value,
                })
        wb.close()
        return remarks
    except Exception as e:
        raise

# AI Agent setup
support_agent = None


def prompt_builder(search_description: str, path: str, sheet_name: str):

    
    
    description_list = get_descriptions_with_index(path, sheet_name)
    prompt = f"""**SHEET DATA:**
                {description_list}

                **SEARCH TEXT:** 
                {search_description}""" + additional_prompt 
    
    return prompt

async def get_llm_result(search_description, path: str, sheet_name: str):
    try:
        if not support_agent:
            raise Exception("AI agent not initialized")
        prompt = prompt_builder(search_description, path, sheet_name)
        response = support_agent.run(prompt)
        return (
            response.content.found_descriptions_list,
            response.content.not_found_descriptions_list,
            response.content.relevant_indexes,
            response.content.updated_quantity,
            response.content.dates,
            response.content.conclution,
        )
    except Exception as e:
        raise

async def updated_quantity_in_sheet(file_path, description: str, sheet_name: str, name: str = "User", location: str = "Home"):
    try:
        (found_descriptions_list, _, relevant_indexes, updated_quantity, dates, conclution) = await get_llm_result(description, file_path, sheet_name)
        if not dates:
            dates = [datetime.date.today()] * len(found_descriptions_list)
        elif len(dates) > 0:
            dates = [datetime.datetime.strptime(date, "%d-%m-%Y").date() for date in dates]
        for found_description, row_index, qty, date in zip(found_descriptions_list, relevant_indexes, updated_quantity, dates):
            col_index = get_date_column(file_path, sheet_name, date)
            if not col_index:
                raise ValueError(f"Could not find today's date in sheet: {sheet_name}")
            update_sheet(file_path=file_path, sheet_name=sheet_name, row_index=row_index, column_index=col_index, value=qty)
            put_logs_in_file(file_path=file_path, sheet_name=sheet_name, description=description,
                              found_description=found_description, row_index=row_index, column_index=col_index,
                              value=qty, name=name, location=location, remark=conclution)
        return True, conclution
    except Exception as e:
        raise

# FastAPI app setup
def create_fastapi_app():
    global fastapi_app
    fastapi_app = FastAPI()
    fastapi_app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
        expose_headers=["*"],
        max_age=600,
    )

    @fastapi_app.options("/get_credentials", response_model=dict)
    async def options_get_credentials():
        return {}

    @fastapi_app.get("/get_credentials")
    async def get_credentials():
        return {
            "GROQ_API_KEY": app_config.get('GROQ_API_KEY'),
            "AVAILABLE_SHEETS": get_available_sheets(app_config.get('EXCEL_FILE_PATH')),
            "AUTHRISED_USERS": app_config.get('ALLOWED_USERS')
        }

    @fastapi_app.options("/process", response_model=dict)
    async def options_process():
        return {}

    @fastapi_app.post("/process")
    async def process_data(transcription: str, sheet_name: Optional[str] = "", name: Optional[str] = "", location: Optional[str] = ""):
        try:
            output = await updated_quantity_in_sheet(app_config.get('EXCEL_FILE_PATH'), transcription, sheet_name, name, location)
            return {"conclution": output[1]}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @fastapi_app.options("/get_history", response_model=dict)
    async def options_get_history():
        return {}

    @fastapi_app.post("/get_history")
    async def get_history_data(name: Optional[str] = "", location: Optional[str] = ""):
        return get_history(app_config.get('EXCEL_FILE_PATH'), name, location)

# def run_server():
#     if fastapi_app:
#         uvicorn.run(fastapi_app, host="0.0.0.0", port=8000, log_level="error")

# ======================== SERVER MANAGER (NO THREADS) ========================

class ServerManager(QObject):
    """Manages server and ngrok tunnel within the same process using threads"""
    status_changed = pyqtSignal(str, str)  # status, message
    url_ready = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.server = None
        self.server_thread = None
        self.tunnel = None
        self.is_running = False
        self.auth_token = None
    
    def cleanup_processes_sync(self):
        """Synchronous cleanup of existing ngrok processes"""
        try:
            # Kill ngrok processes
            for proc in psutil.process_iter(['pid', 'name']):
                try:
                    if proc.info['name'] and 'ngrok' in proc.info['name'].lower():
                        proc.kill()
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    continue
            # Clean ngrok
            try:
                ngrok.kill()
            except:
                pass
        except Exception as e:
            print(f"Cleanup error: {e}")
    
    def start_server_and_tunnel(self, auth_token):
        """Start server and tunnel in a thread"""
        if self.is_running:
            return
        self.auth_token = auth_token
        self.is_running = True
        self.status_changed.emit("starting", "Cleaning up existing processes...")
        
        # Perform cleanup
        self.cleanup_processes_sync()
        
        self.status_changed.emit("starting", "Starting server...")
        
        # Configure and start server in a thread
        try:
            config = uvicorn.Config(fastapi_app, host="0.0.0.0", port=8000, log_level="error")
            self.server = uvicorn.Server(config)
            self.server_thread = threading.Thread(target=self.server.run, daemon=True)
            self.server_thread.start()
            
            # Wait briefly before starting tunnel
            QTimer.singleShot(3000, self.create_tunnel)
        except Exception as e:
            self.is_running = False
            self.error_occurred.emit(f"Failed to start server: {str(e)}")
    
    def create_tunnel(self):
        """Create ngrok tunnel"""
        try:
            self.status_changed.emit("starting", "Starting tunnel...")
            if self.auth_token:
                ngrok.set_auth_token(self.auth_token)
            self.tunnel = ngrok.connect(8000, "http")
            public_url = self.tunnel.public_url
            self.status_changed.emit("running", "Server & Tunnel Running")
            self.url_ready.emit(public_url)
        except Exception as e:
            self.is_running = False
            self.error_occurred.emit(f"Failed to create tunnel: {str(e)}")
    
    def stop_server_and_tunnel(self):
        """Stop server and tunnel"""
        if not self.is_running:
            return
        self.is_running = False
        self.status_changed.emit("stopping", "Stopping server and tunnel...")
        
        # Stop server
        if self.server:
            try:
                self.server.should_exit = True
                if self.server_thread:
                    self.server_thread.join(timeout=5)
                self.server = None
                self.server_thread = None
            except Exception as e:
                print(f"Error stopping server: {e}")
        
        # Stop tunnel
        if self.tunnel:
            try:
                ngrok.disconnect(self.tunnel.public_url)
                self.tunnel = None
            except Exception as e:
                print(f"Error stopping tunnel: {e}")
        
        self.status_changed.emit("stopped", "Server stopped")

# ======================== MAIN APPLICATION ========================

class NgrokTunnelApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.server_manager = ServerManager()
        self.load_configuration()
        self.init_ui()
        self.setup_connections()
        # Register cleanup function
        atexit.register(self.cleanup_on_exit)
        
    def cleanup_on_exit(self):
        """Cleanup function called on exit"""
        if self.server_manager and self.server_manager.is_running:
            self.server_manager.stop_server_and_tunnel()
        
    def setup_connections(self):
        """Setup signal connections"""
        self.server_manager.status_changed.connect(self.on_status_changed)
        self.server_manager.url_ready.connect(self.on_url_ready)
        self.server_manager.error_occurred.connect(self.on_error_occurred)
        
    def load_configuration(self):
        global app_config
        app_config = load_config()
        
        # Set defaults if not configured
        if not app_config:
            app_config = {
                'NGROK_AUTH_TOKEN': '',
                'GROQ_API_KEY': '',
                'EXCEL_FILE_PATH': '',
                'ALLOWED_USERS': [{"name": "DEFAULT", "location": "HOME"}]
            }
        
        # Initialize server components
        create_fastapi_app()
        initialize_ai_agent()
    
    def init_ui(self):
        self.setWindowTitle("SMART DPR")
        self.setGeometry(100, 100, 1000, 800)
        
        # Set dark theme
        self.setStyleSheet("""
                QWidget {
                    background-color: #f5f5f5;
                    font-family: -apple-system, BlinkMacSystemFont, 'Helvetica Neue', Arial, sans-serif;
                    font-size: 10pt;
                    color: #333333;
                }

                QTabWidget::pane {
                    border: 1px solid #c0c0c0;
                    background-color: white;
                    border-radius: 5px;
                    color: #333333;
                }

                QTabBar::tab {
                    background-color: #e1e1e1;
                    border: 1px solid #c0c0c0;
                    padding: 8px 16px;
                    margin-right: 2px;
                    border-top-left-radius: 5px;
                    border-top-right-radius: 5px;
                    color: #333333;
                }

                QTabBar::tab:selected {
                    background-color: white;
                    border-bottom: 1px solid white;
                    color: #333333;
                }

                QTabBar::tab:hover {
                    background-color: #d1d1d1;
                    color: #333333;
                }

                QPushButton {
                    background-color: #0078d4;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                    min-width: 80px;
                }

                QPushButton:hover {
                    background-color: #106ebe;
                }

                QPushButton:pressed {
                    background-color: #005a9e;
                }

                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }

                QPushButton.danger {
                    background-color: #d13438;
                    color: white;
                }

                QPushButton.danger:hover {
                    background-color: #b02a2f;
                    color: white;
                }

                QPushButton.success {
                    background-color: #107c10;
                }

                QPushButton.success:hover {
                    background-color: #0e6e0e;
                }

                QPushButton.secondary {
                    background-color: #6c757d;
                }

                QPushButton.secondary:hover {
                    background-color: #5a6268;
                }

                QPushButton.link {
                    background-color: transparent;
                    color: #0078d4;
                    border: 1px solid #0078d4;
                    text-decoration: underline;
                    padding: 4px 8px;
                    min-width: 60px;
                }

                QPushButton.link:hover {
                    background-color: #e7f3ff;
                    color: #005a9e;
                }

                QLineEdit {
                    border: 1px solid #d1d1d1;
                    border-radius: 4px;
                    padding: 8px;
                    background-color: white;
                    font-size: 10pt;
                    color: #333333;
                }

                QLineEdit:focus {
                    border: 2px solid #0078d4;
                    color: #333333;
                }

                QLineEdit:read-only {
                    background-color: #f8f9fa;
                    color: #6c757d;
                }

                QTextEdit {
                    border: 1px solid #d1d1d1;
                    border-radius: 4px;
                    background-color: white;
                    padding: 8px;
                    font-family: 'Courier New', 'Monaco', 'Menlo', monospace;
                    color: #333333;
                }

                QLabel {
                    color: #333333;
                    font-weight: 500;
                    background-color: transparent;
                }

                QGroupBox {
                    font-weight: bold;
                    border: 1px solid #d1d1d1;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding-top: 10px;
                    color: #333333;
                    background-color: transparent;
                }

                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px 0 5px;
                    color: #333333;
                    background-color: transparent;
                }

                QTableWidget {
                    border: 1px solid #d1d1d1;
                    border-radius: 4px;
                    background-color: white;
                    gridline-color: #e1e1e1;
                    selection-background-color: #0078d4;
                    color: #333333;
                }

                QTableWidget::item {
                    padding: 8px;
                    color: #333333;
                    background-color: white;
                }

                QTableWidget::item:selected {
                    background-color: #0078d4;
                    color: white;
                }

                QHeaderView::section {
                    background-color: #f8f9fa;
                    padding: 8px;
                    border: 1px solid #d1d1d1;
                    font-weight: bold;
                    color: #333333;
                }

                QProgressBar {
                    border: 1px solid #d1d1d1;
                    border-radius: 4px;
                    background-color: #f8f9fa;
                    text-align: center;
                    color: #333333;
                }

                QProgressBar::chunk {
                    background-color: #107c10;
                    border-radius: 3px;
                }

                QFrame.separator {
                    border: none;
                    border-top: 1px solid #d1d1d1;
                    margin: 10px 0;
                }

                QScrollArea {
                    border: none;
                    background-color: transparent;
                }

                QScrollArea > QWidget > QWidget {
                    background-color: transparent;
                }
        """)
        
        # Create main widget and tab widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # Create tabs
        self.create_configuration_tab()
        self.create_connect_devices_tab()
        self.create_support_tab()
        
    def create_configuration_tab(self):
        """Create the Configuration tab"""
        config_widget = QWidget()
        config_layout = QVBoxLayout()
        config_widget.setLayout(config_layout)
        
        # Add some spacing at the top
        config_layout.addSpacing(10)
        
        # Configuration title
        # title_label = QLabel("Configuration")
        # title_label.setAlignment(Qt.AlignCenter)
        # title_label.setFont(QFont("Arial", 16, QFont.Bold))
        # title_label.setStyleSheet("color: #333333; background-color: transparent;")
        # config_layout.addWidget(title_label)
        
        # API Configuration Group
        api_group = QGroupBox("API Configuration")
        api_layout = QGridLayout()
        
        # Ngrok Auth Token
        api_layout.addWidget(QLabel("Ngrok Auth Token:"))
        self.ngrok_token_input = QLineEdit()
        self.ngrok_token_input.setEchoMode(QLineEdit.Password)
        self.ngrok_token_input.setText(app_config.get('NGROK_AUTH_TOKEN', ''))
        self.ngrok_token_input.setPlaceholderText("Enter your ngrok auth token")
        api_layout.addWidget(self.ngrok_token_input, 0, 1)
        
        get_token_btn = QPushButton("Get Token")
        get_token_btn.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://dashboard.ngrok.com/get-started/your-authtoken")))
        get_token_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        api_layout.addWidget(get_token_btn, 0, 2)
        
        # GROQ API Key
        api_layout.addWidget(QLabel("GROQ API Key:"), 1, 0)
        self.groq_key_input = QLineEdit()
        self.groq_key_input.setEchoMode(QLineEdit.Password)
        self.groq_key_input.setText(app_config.get('GROQ_API_KEY', ''))
        self.groq_key_input.setPlaceholderText("Enter your groq api key")
        api_layout.addWidget(self.groq_key_input, 1, 1)
        
        get_api_key_btn = QPushButton("Get API Key")
        get_api_key_btn.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://console.groq.com/keys")))
        get_api_key_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        api_layout.addWidget(get_api_key_btn, 1, 2)
        
        api_group.setLayout(api_layout)
        config_layout.addWidget(api_group)
        
        config_layout.addSpacing(20)
        
        # File Configuration Group
        file_group = QGroupBox("File Configuration")
        file_layout = QGridLayout()
        
        file_layout.addWidget(QLabel("Excel File Path:"), 0, 0)
        self.excel_path_input = QLineEdit()
        self.excel_path_input.setReadOnly(True)  # Make the field read-only
        self.excel_path_input.setText(app_config.get('EXCEL_FILE_PATH', ''))
        self.excel_path_input.setPlaceholderText("Click 'Select Excel File' to choose a file")
        self.excel_path_input.setStyleSheet("""
            QLineEdit:read-only {
                background-color: #f0f0f0;
                border: 1px solid #d1d1d1;
                color: #333333;
            }
        """)
        file_layout.addWidget(self.excel_path_input, 0, 1)
        
        select_file_btn = QPushButton("Select Excel File")
        select_file_btn.clicked.connect(self.browse_excel_file)
        select_file_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        file_layout.addWidget(select_file_btn, 0, 2)
        
        file_group.setLayout(file_layout)
        config_layout.addWidget(file_group)
        
        config_layout.addSpacing(20)
        
        # Authorized Users Management Group
        users_group = QGroupBox("Authorized Users Management")
        users_layout = QVBoxLayout()
        
        # Add user section
        add_user_layout = QHBoxLayout()
        
        name_label = QLabel("Name:")
        name_label.setFixedWidth(50)
        add_user_layout.addWidget(name_label)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter user name")
        add_user_layout.addWidget(self.name_input)
        
        location_label = QLabel("Location/Area:")
        location_label.setFixedWidth(60)
        add_user_layout.addWidget(location_label)
        
        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("Enter user location")
        add_user_layout.addWidget(self.location_input)
        
        add_user_btn = QPushButton("Add User")
        add_user_btn.clicked.connect(self.add_user)
        add_user_btn.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        add_user_layout.addWidget(add_user_btn)
        
        users_layout.addLayout(add_user_layout)
        
        users_layout.addSpacing(10)
        
        # Users table
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(3)
        self.users_table.setHorizontalHeaderLabels(["User Name", "Location", "Select"])
        
        # Disable selection highlighting and focus
        self.users_table.setSelectionMode(QTableWidget.NoSelection)
        self.users_table.setFocusPolicy(Qt.NoFocus)
        self.users_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # Set column widths
        header = self.users_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        self.users_table.setColumnWidth(2, 80)
        
        # Styling to remove selection highlight
        self.users_table.setStyleSheet("""
            QTableWidget {
                selection-background-color: transparent;
                selection-color: black;
                outline: none;
            }
            QTableWidget::item:selected {
                background: transparent;
                color: black;
            }
        """)
        
        self.populate_users_table()
        users_layout.addWidget(self.users_table)
        
        # User management buttons
        user_buttons_layout = QHBoxLayout()
        
        delete_selected_btn = QPushButton("Delete Selected")
        delete_selected_btn.clicked.connect(self.delete_selected_users)
        delete_selected_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        user_buttons_layout.addWidget(delete_selected_btn)
        
        clear_all_btn = QPushButton("Clear All")
        clear_all_btn.clicked.connect(self.clear_all_users)
        clear_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #666666;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #555555;
            }
        """)
        user_buttons_layout.addWidget(clear_all_btn)
        
        user_buttons_layout.addStretch()
        users_layout.addLayout(user_buttons_layout)
        
        users_group.setLayout(users_layout)
        config_layout.addWidget(users_group)
        
        config_layout.addSpacing(30)
        
        # Save Configuration Button
        save_config_btn = QPushButton("Save Configuration")
        save_config_btn.clicked.connect(self.save_current_config)
        save_config_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                padding: 15px 30px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        config_layout.addWidget(save_config_btn, alignment=Qt.AlignCenter)
        
        config_layout.addStretch()
        
        self.tab_widget.addTab(config_widget, "Configuration")
    
    def create_connect_devices_tab(self):
        """Create the Connect Devices tab"""
        connect_widget = QWidget()
        connect_layout = QVBoxLayout()
        connect_widget.setLayout(connect_layout)
        
        # Add some spacing at the top
        connect_layout.addSpacing(50)
        
        # Connect Devices title
        title_label = QLabel("Connect Devices")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 30px;")
        connect_layout.addWidget(title_label)
        
        connect_layout.addSpacing(50)
        
        # Server status section
        status_group = QGroupBox("Server Status")
        status_layout = QVBoxLayout()
        
        # Status display
        self.status_label = QLabel("Status: Not connected")
        self.status_label.setStyleSheet("font-weight: bold; color: #cccccc; font-size: 16px; padding: 10px;")
        self.status_label.setAlignment(Qt.AlignCenter)
        status_layout.addWidget(self.status_label)
        
        # URL display with copy button
        url_container = QWidget()
        url_layout = QHBoxLayout()
        url_layout.setContentsMargins(10, 5, 10, 5)
        
        # Add APP PASSWORD label
        password_label = QLabel("APP PASSWORD:")
        password_label.setStyleSheet("font-weight: bold; color: #333333; font-size: 14px; padding: 5px;")
        url_layout.addWidget(password_label)
        
        # Password value
        self.url_label = QLabel("")
        self.url_label.setStyleSheet("""
            QLabel {
                color: #1976d2;
                font-size: 14px;
                font-family: monospace;
                background-color: #f0f0f0;
                padding: 5px 10px;
                border-radius: 3px;
                border: 1px solid #dddddd;
                min-width: 100px;
            }
        """)
        url_layout.addWidget(self.url_label, 1)
        
        # Add copy button
        self.copy_button = QPushButton("Copy")
        self.copy_button.setStyleSheet("""
            QPushButton {
                background-color: #1976d2;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
                font-size: 12px;
                margin-left: 10px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        self.copy_button.clicked.connect(self.copy_password_to_clipboard)
        url_layout.addWidget(self.copy_button)
        
        url_container.setLayout(url_layout)
        url_container.hide()
        self.url_container = url_container  # Store reference to show/hide with the label
        status_layout.addWidget(url_container)
        
        status_group.setLayout(status_layout)
        connect_layout.addWidget(status_group)
        
        connect_layout.addSpacing(50)
        
        # Control buttons
        control_layout = QHBoxLayout()
        control_layout.addStretch()
        
        # Start/Stop button
        self.toggle_button = QPushButton("Start Server")
        self.toggle_button.clicked.connect(self.toggle_server)
        self.toggle_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 20px 40px;
                font-size: 18px;
                font-weight: bold;
                border-radius: 10px;
                min-width: 250px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #999999;
            }
        """)
        control_layout.addWidget(self.toggle_button)
        control_layout.addStretch()
        
        connect_layout.addLayout(control_layout)
        connect_layout.addStretch()
        
        self.tab_widget.addTab(connect_widget, "Connect Devices")
    
    def create_support_tab(self):
        """Create the Support tab"""
        support_widget = QWidget()
        support_layout = QVBoxLayout()
        support_widget.setLayout(support_layout)
        
        # Add some spacing at the top
        support_layout.addSpacing(30)
        
        # Support title
        title_label = QLabel("SMART DPR")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 24px; 
            font-weight: bold; 
            margin-bottom: 20px;
            color: #1976d2;
        """)
        support_layout.addWidget(title_label)
        
        # Main content
        content = QTextEdit()
        content.setReadOnly(True)
        content.setStyleSheet("""
            QTextEdit {
                background-color: #2b2b2b;
                border: 1px solid #444444;
                border-radius: 5px;
                padding: 20px;
                color: #e0e0e0;
                font-size: 14px;
                line-height: 1.6;
            }
            a {
                color: #64b5f6;
                text-decoration: none;
            }
            a:hover {
                text-decoration: underline;
            }
        """)
        
        # Formatted support content
        support_content = """
        <h3 style='color: #64b5f6; margin-top: 0;'>About SMART DPR</h3>
        <p>SMART DPR is currently in the MVP phase and open-sourced. We welcome feedback, bug reports, and your custom feature requests.</p>
        
        <p>If you face any issues or need modifications, feel free to contact our support team. We're actively listening and here to help you make SMART DPR better. Our team will assist you with installation and setup in case of any problems.</p>
        
        <h3 style='color: #64b5f6;'>Have Feedback?</h3>
        <p>This is our first initiative, and we're building it with you and for you. We're constantly improving and would love to hear from you.</p>
        
        <p>✅ <strong>Tell us what challenges you face in real estate.</strong><br>
        We're actively looking to solve real-world real estate problems using Agentic AI—whether it's automation, insights, sales support, or anything else.</p>
        
        <h3 style='color: #64b5f6;'>📩 Contact Support</h3>
        <p>Share your thoughts, frustrations, or feature requests:</p>
        <p>
            <strong>Email:</strong> <a href='mailto:innovate@xaneur.com'>innovate@xaneur.com</a><br>
            <strong>Phone:</strong> <a href='tel:+919510595426'>+91 95105 95426</a>
        </p>
        
        <div style='margin-top: 30px; padding-top: 15px; border-top: 1px solid #444; font-size: 12px; color: #888;'>

        </div>
        """
        
        content.setHtml(support_content)
        support_layout.addWidget(content, 1)  # The '1' makes it take available space
        
        # Add some spacing at the bottom
        support_layout.addSpacing(20)
        
        self.tab_widget.addTab(support_widget, "Support")
    
    def populate_users_table(self):
        """Populate the users table with current users"""
        users = app_config.get('ALLOWED_USERS', [])
        self.users_table.setRowCount(len(users))
        
        for row, user in enumerate(users):
            # User name
            name_item = QTableWidgetItem(user.get('name', ''))
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.users_table.setItem(row, 0, name_item)
            
            # Location
            location_item = QTableWidgetItem(user.get('location', ''))
            location_item.setFlags(location_item.flags() & ~Qt.ItemIsEditable)
            self.users_table.setItem(row, 1, location_item)
            
            # Checkbox
            checkbox = QCheckBox()
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout()
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            checkbox_widget.setLayout(checkbox_layout)
            self.users_table.setCellWidget(row, 2, checkbox_widget)
    
    def add_user(self):
        """Add a new user"""
        name = self.name_input.text().strip()
        location = self.location_input.text().strip()
        
        if not name or not location:
            QMessageBox.warning(self, "Validation Error", "Please enter both name and location.")
            return
        
        # Check if username already exists (case-insensitive check)
        users = app_config.get('ALLOWED_USERS', [])
        for user in users:
            if user.get('name', '').lower() == name.lower():
                QMessageBox.warning(
                    self, 
                    "Username Exists", 
                    f"A user with the name '{name}' already exists.\n\n"
                    "Please choose a different username."
                )
                return
        
        # Show success message
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("User Added")
        msg.setText(f"User '{name}' from '{location}' added successfully")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #2b2b2b;
                color: #ffffff;
            }
            QLabel {
                color: #ffffff;
                font-size: 14px;
            }
            QPushButton {
                background-color: #1976d2;
                color: white;
                padding: 8px 20px;
                border-radius: 4px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
        """)
        msg.exec_()
        
        # Add new user to the list (but don't save to config yet)
        users.append({"name": name, "location": location})
        app_config['ALLOWED_USERS'] = users
        
        # Clear inputs
        self.name_input.clear()
        self.location_input.clear()
        
        # Refresh table
        self.populate_users_table()
    
    def delete_selected_users(self):
        """Delete selected users"""
        users_to_delete = []
        
        for row in range(self.users_table.rowCount()):
            checkbox_widget = self.users_table.cellWidget(row, 2)
            checkbox = checkbox_widget.findChild(QCheckBox)
            if checkbox and checkbox.isChecked():
                name = self.users_table.item(row, 0).text()
                location = self.users_table.item(row, 1).text()
                users_to_delete.append({"name": name, "location": location})
        
        if not users_to_delete:
            QMessageBox.information(self, "No Selection", "Please select users to delete.")
            return
        
        reply = QMessageBox.question(
            self, 'Confirm Delete',
            f'Are you sure you want to delete {len(users_to_delete)} user(s)?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            users = app_config.get('ALLOWED_USERS', [])
            for user_to_delete in users_to_delete:
                users = [u for u in users if not (u.get('name') == user_to_delete['name'] and u.get('location') == user_to_delete['location'])]
            
            app_config['ALLOWED_USERS'] = users
            self.populate_users_table()
            self.save_current_config()
    
    def clear_all_users(self):
        """Clear all users"""
        reply = QMessageBox.question(
            self, 'Confirm Clear All',
            'Are you sure you want to delete all users?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            app_config['ALLOWED_USERS'] = []
            self.populate_users_table()
            self.save_current_config()
    
    def browse_excel_file(self):
        """Browse for Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.excel_path_input.setText(file_path)
            self.save_current_config()
    
    def save_current_config(self):
        """Save current configuration from UI inputs"""
        try:
            # Update global config
            app_config.update({
                'NGROK_AUTH_TOKEN': self.ngrok_token_input.text().strip(),
                'GROQ_API_KEY': self.groq_key_input.text().strip(),
                'EXCEL_FILE_PATH': self.excel_path_input.text().strip(),
                'ALLOWED_USERS': app_config.get('ALLOWED_USERS', [])
            })
            
            # Save to file
            save_config(app_config)
            
            # Reinitialize components
            create_fastapi_app()
            initialize_ai_agent()
            
            # Show success message
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Success")
            msg.setText("Configuration saved successfully!")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.setStyleSheet("""
                QMessageBox {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                    font-size: 14px;
                }
                QPushButton {
                    background-color: #1976d2;
                    color: white;
                    padding: 8px 20px;
                    border-radius: 4px;
                    min-width: 80px;
                }
                QPushButton:hover {
                    background-color: #1565c0;
                }
            """)
            msg.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Configuration error: {str(e)}")
    
    def validate_configuration(self):
        """Validate current configuration"""
        errors = []
        
        if not app_config.get('NGROK_AUTH_TOKEN', '').strip():
            errors.append("Ngrok Auth Token is required")
        
        if not app_config.get('GROQ_API_KEY', '').strip():
            errors.append("Groq API Key is required")
        
        excel_path = app_config.get('EXCEL_FILE_PATH', '').strip()
        if not excel_path:
            errors.append("Excel File Path is required")
        elif not os.path.exists(excel_path):
            errors.append("Excel file does not exist")
        
        if not app_config.get('ALLOWED_USERS'):
            errors.append("At least one user must be configured")
        
        return errors
    
    def toggle_server(self):
        if self.toggle_button.text() == "Start Server":
            self.start_server()
        else:
            self.stop_server()
    
    def start_server(self):
        # Validate configuration
        errors = self.validate_configuration()
        if errors:
            QMessageBox.warning(
                self, 
                "Configuration Error", 
                "Please fix the following configuration issues:\n\n" + "\n".join(f"• {error}" for error in errors)
            )
            return
        
        self.toggle_button.setEnabled(False)
        self.toggle_button.setText("Starting...")
        
        # Start server using timer-based approach (no threads)
        self.server_manager.start_server_and_tunnel(app_config.get('NGROK_AUTH_TOKEN'))
    
    def stop_server(self):
        self.toggle_button.setEnabled(False)
        self.toggle_button.setText("Stopping...")
        
        # Stop server synchronously (no threads)
        self.server_manager.stop_server_and_tunnel()
    
    def on_status_changed(self, status, message):
        """Handle status changes from server manager"""
        self.status_label.setText(f"Status: {message}")
        
        if status == "starting":
            self.toggle_button.setEnabled(False)
        elif status == "started":
            self.toggle_button.setEnabled(True)
            self.toggle_button.setText("Stop Server")
        elif status in ["stopped", "error"]:
            self.toggle_button.setEnabled(True)
            self.toggle_button.setText("Start Server")
            self.toggle_button.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 20px 40px;
                    font-size: 18px;
                    font-weight: bold;
                    border-radius: 10px;
                    min-width: 250px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:disabled {
                    background-color: #666666;
                    color: #999999;
                }
            """)
            if hasattr(self, 'url_container'):
                self.url_container.hide()
    
    def copy_password_to_clipboard(self):
        """Copy the password to clipboard"""
        if hasattr(self, 'current_subdomain'):
            clipboard = QApplication.clipboard()
            clipboard.setText(self.current_subdomain)
            
            # Show a brief tooltip or change button text to confirm copy
            self.copy_button.setText("Copied!")
            QTimer.singleShot(2000, lambda: self.copy_button.setText("Copy"))
    
    def on_url_ready(self, url):
        """Handle successful tunnel creation"""
        subdomain = re.sub(r'^https?://([^.]+).*$', r'\1', url)
        
        self.toggle_button.setEnabled(True)
        self.toggle_button.setText("Stop Server")
        self.toggle_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                padding: 20px 40px;
                font-size: 18px;
                font-weight: bold;
                border-radius: 10px;
                min-width: 250px;
                margin-top: 20px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #999999;
            }
        """)
        
        # Store the subdomain without extra spaces
        self.current_subdomain = subdomain
        # Display with spaces for better UI
        self.url_label.setText(subdomain)
        self.url_container.show()
        
        # Copy URL to clipboard
        self.copy_password_to_clipboard()
        
        QMessageBox.information(
            self, 
            "Stop Server Started", 
            f"Server and tunnel started successfully!\n\n"
            f"Public URL: {url}\n"
        )
    
    def on_error_occurred(self, error_message):
        """Handle errors from server manager"""
        self.toggle_button.setEnabled(True)
        self.toggle_button.setText("Start Server")
        self.toggle_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 20px 40px;
                font-size: 18px;
                font-weight: bold;
                border-radius: 10px;
                min-width: 250px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #999999;
            }
        """)
        
        self.url_label.hide()
        
        QMessageBox.critical(self, "Error", f"Failed to start server/tunnel:\n\n{error_message}")
    
    def closeEvent(self, event):
        """Handle application close event safely"""
        try:
            reply = QMessageBox.question(
                self, 'Confirm Exit',
                'Are you sure you want to exit?\n\nThis will stop the server and close any active tunnels.',
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Stop server safely before closing
                if self.server_manager and self.server_manager.is_running:
                    self.server_manager.stop_server_and_tunnel()
                
                event.accept()
            else:
                event.ignore()
                
        except Exception as e:
            # If there's any error during shutdown, still allow exit
            print(f"Error during shutdown: {e}")
            event.accept()

def main():
    # Handle Ctrl+C gracefully
    signal.signal(signal.SIGINT, signal.SIG_DFL)
    
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("SMART DPR")
    app.setApplicationVersion("1.0")
    
    # Create and show main window
    window = NgrokTunnelApp()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()


